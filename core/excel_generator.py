"""
excel_generator.py – Generates a formatted Excel P&L report from parsed MIS data.

Layout mirrors the manual MIS reference exactly:
  P&L sheet  – current-month columns | separator | YTD columns
  Drs_Crs    – Sundry Debtors & Creditors summary
"""

import os
import uuid
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Excel number format: positive = comma-formatted; negative = same; zero = dash
NUM_FMT = '#,##0.00;-#,##0.00;"-"'
PCT_FMT = '0.00%;-0.00%;"-"'

DARK_BLUE = "002060"
GREEN     = "385D3A"
LIGHT_BLUE = "DCE6F1"
YELLOW    = "FFFF99"
WHITE_FONT = Font(color="FFFFFF", bold=True)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe(v: Any) -> float:
    """Return v as float; None/empty → 0.0."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _pct_str(numerator: float, denominator: float) -> str:
    if denominator == 0:
        return "0.00%"
    return f"{(numerator / denominator) * 100:.2f}%"


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_excel_report(report_data: dict) -> str:
    temp_dir = os.path.join(os.getcwd(), "temp_reports")
    os.makedirs(temp_dir, exist_ok=True)
    filename = f"report_{uuid.uuid4().hex}.xlsx"
    filepath = os.path.join(temp_dir, filename)

    company_name  = report_data.get("company_name",  "Company")
    report_month  = report_data.get("report_month",  "Month Year")
    ytd_range     = report_data.get("ytd_range",     "(YTD)")
    vert_types    = report_data.get("vertical_types", {})

    # ------------------------------------------------------------------ #
    # 1.  Build ordered vertical list                                     #
    # ------------------------------------------------------------------ #
    # Use the natural vertical order emitted by the parser (first-seen in List of Ledgers).
    # This preserves the exact column sequence for any company without hardcoding names.
    all_verts: list[str] = list(report_data["verticals"].keys())
    vertical_order: list[str] = report_data.get("vertical_order", all_verts)

    # Derive vertical_types from data if the parser didn't supply it
    if not vert_types:
        for v, d in report_data["verticals"].items():
            v_lower = v.lower()
            has_rev = any([d.get("revenue"), d.get("cogs"), d.get("direct_expenses"),
                           d.get("ytd_revenue"), d.get("ytd_cogs")])
            if "share" in v_lower and "trading" in v_lower:
                vert_types[v] = "share_trading"
            elif not has_rev:
                vert_types[v] = "cost_center"
            else:
                vert_types[v] = "revenue"

    def _order_key(name: str) -> int:
        """Sort key based on the parser's natural vertical_order."""
        try:
            return vertical_order.index(name)
        except ValueError:
            return len(vertical_order)  # unknown verticals go to the end

    rev_verts     = sorted([v for v in all_verts if vert_types.get(v) == "revenue"],     key=_order_key)
    cost_centers  = sorted([v for v in all_verts if vert_types.get(v) == "cost_center"], key=_order_key)
    share_trading = sorted([v for v in all_verts if vert_types.get(v) == "share_trading"], key=_order_key)

    # Drop Unallocated if empty
    for bucket in [rev_verts, cost_centers, share_trading]:
        if "Unallocated" in bucket:
            d = report_data["verticals"]["Unallocated"]
            if not any([d.get("revenue"), d.get("cogs"), d.get("indirect_expenses"),
                        d.get("indirect_income")]):
                bucket.remove("Unallocated")

    # Full ordered list for table iteration
    ordered_verts = rev_verts + cost_centers + share_trading

    # ------------------------------------------------------------------ #
    # 2.  Gather all ledger names for row labels                          #
    # ------------------------------------------------------------------ #
    all_direct_ledgers: list[str] = []
    all_income_ledgers: list[str] = []
    all_indirect_ledgers: list[str] = []

    _direct_seen:   set[str] = set()
    _income_seen:   set[str] = set()
    _indirect_seen: set[str] = set()

    for v in ordered_verts:
        vd = report_data["verticals"][v]
        for k in list(vd.get("direct_breakdown", {}).keys()) + list(vd.get("ytd_direct_breakdown", {}).keys()):
            if k not in _direct_seen:
                _direct_seen.add(k); all_direct_ledgers.append(k)
        for k in list(vd.get("income_breakdown", {}).keys()) + list(vd.get("ytd_income_breakdown", {}).keys()):
            if k not in _income_seen:
                _income_seen.add(k); all_income_ledgers.append(k)
        for k in list(vd.get("indirect_breakdown", {}).keys()) + list(vd.get("ytd_indirect_breakdown", {}).keys()):
            if k not in _indirect_seen:
                _indirect_seen.add(k); all_indirect_ledgers.append(k)

    # Remove Depreciation rows – they appear in their own dedicated row
    for depr_name in ["Depreciation & amortization", "Depreciation"]:
        if depr_name in all_indirect_ledgers:
            all_indirect_ledgers.remove(depr_name)

    # ------------------------------------------------------------------ #
    # 3.  P&L row labels                                                  #
    # ------------------------------------------------------------------ #
    pl_rows: list[str] = ["Sales", "Less: COGS"]

    if all_direct_ledgers:
        pl_rows += ["3. Direct Expense"] + all_direct_ledgers

    pl_rows += ["Gross margin", "Gross margin %", "Gross margin % (previous month)", ""]

    if all_income_ledgers:
        pl_rows += ["Indirect Income"] + all_income_ledgers
    else:
        pl_rows += ["Indirect Income"]

    pl_rows += [
        "Depreciation & amortization", "Net income", "Net allocable income", "",
        "6. Indirect Expense"
    ]
    pl_rows += all_indirect_ledgers
    pl_rows += ["Indirect costs", "", "Allocation of expenses:"]
    pl_rows += cost_centers
    pl_rows += [
        "Total indirect costs", "",
        "Profit/ (loss) before tax", "Net margin %", "Net margin % (previous month)", "", "",
        "Profit/(loss) as per tally", "Difference",
    ]

    LABEL_IDX: dict[str, int] = {r: i for i, r in enumerate(pl_rows)}

    # ------------------------------------------------------------------ #
    # 4.  Column builder functions                                         #
    # ------------------------------------------------------------------ #
    def _build_col(vd: dict, is_ytd: bool) -> tuple[list, dict]:
        """
        Returns (col_values, extra_nums) where:
          col_values – list aligned to pl_rows, values are float|str
          extra_nums – dict of {row_label: float} for percentage rows (as strings)
        """
        pfx = "ytd_" if is_ytd else ""
        sales   = _safe(vd.get(f"{pfx}revenue"))
        cogs    = _safe(vd.get(f"{pfx}cogs"))
        dir_exp = _safe(vd.get(f"{pfx}direct_expenses"))
        gross   = sales - cogs - dir_exp
        ind_inc = _safe(vd.get(f"{pfx}indirect_income"))
        depr    = (
            _safe(vd.get(f"{pfx}indirect_breakdown", {}).get("Depreciation & amortization"))
            + _safe(vd.get(f"{pfx}indirect_breakdown", {}).get("Depreciation"))
        )
        net_inc      = ind_inc
        net_alloc    = 0.0
        net_prof     = gross + ind_inc - depr
        ind_exp      = _safe(vd.get(f"{pfx}indirect_expenses"))
        gross_pct    = _pct_str(gross, sales)
        net_pct_str  = "0.00%"  # placeholder; filled after allocation

        col: list = [0.0] * len(pl_rows)

        def _set(label: str, value: Any) -> None:
            if label in LABEL_IDX:
                col[LABEL_IDX[label]] = value

        _set("Sales",                          sales)
        _set("Less: COGS",                     cogs)
        if all_direct_ledgers:
            _set("3. Direct Expense", dir_exp)
            for l in all_direct_ledgers:
                _set(l, _safe(vd.get(f"{pfx}direct_breakdown", {}).get(l, 0.0)))
        _set("Gross margin",                   gross)
        _set("Gross margin %",                 gross_pct)
        _set("Gross margin % (previous month)", "0.00%")

        if all_income_ledgers:
            _set("Indirect Income", ind_inc)
            for l in all_income_ledgers:
                _set(l, _safe(vd.get(f"{pfx}income_breakdown", {}).get(l, 0.0)))
        else:
            _set("Indirect Income", ind_inc)

        _set("Depreciation & amortization", depr)
        _set("Net income",                  net_inc)
        _set("Net allocable income",        net_alloc)
        _set("6. Indirect Expense",         0.0)   # section header

        tot_ind = 0.0
        for l in all_indirect_ledgers:
            val = _safe(vd.get(f"{pfx}indirect_breakdown", {}).get(l, 0.0))
            _set(l, val)
            tot_ind += val
        _set("Indirect costs", tot_ind)

        _set("Allocation of expenses:", 0.0)   # section header

        tot_alloc = 0.0
        v_name = None  # will be set by the caller loop
        return col, {
            "sales": sales, "cogs": cogs, "dir_exp": dir_exp, "gross": gross,
            "ind_inc": ind_inc, "depr": depr, "net_inc": net_inc, "net_alloc": net_alloc,
            "net_prof": net_prof, "tot_ind": tot_ind,
            "ind_breakdown": vd.get(f"{pfx}indirect_breakdown", {}),
            "dir_breakdown": vd.get(f"{pfx}direct_breakdown", {}),
            "income_breakdown": vd.get(f"{pfx}income_breakdown", {}),
        }

    def _finalise_col(col: list, nums: dict, v_name: str,
                      is_rev: bool, pfx: str) -> list:
        """Fill allocation rows and profit rows (requires knowing the vertical name)."""
        tot_alloc = 0.0
        for cc in cost_centers:
            cc_data = report_data["verticals"][cc]
            cc_costs = sum(cc_data.get(f"{pfx}indirect_breakdown", {}).values())
            alloc_val = 0.0
            if is_rev and len(rev_verts) > 0:
                alloc_val = cc_costs / len(rev_verts)
            elif v_name == cc:
                alloc_val = -cc_costs
            if cc in LABEL_IDX:
                col[LABEL_IDX[cc]] = alloc_val
            tot_alloc += alloc_val

        if "Total indirect costs" in LABEL_IDX:
            col[LABEL_IDX["Total indirect costs"]] = nums["tot_ind"] + tot_alloc

        profit = nums["net_prof"] - nums["tot_ind"] - tot_alloc
        net_pct = _pct_str(profit, nums["sales"])
        if "Profit/ (loss) before tax" in LABEL_IDX:
            col[LABEL_IDX["Profit/ (loss) before tax"]] = profit
        if "Net margin %" in LABEL_IDX:
            col[LABEL_IDX["Net margin %"]] = net_pct
        if "Net margin % (previous month)" in LABEL_IDX:
            col[LABEL_IDX["Net margin % (previous month)"]] = "0.00%"
        if "Profit/(loss) as per tally" in LABEL_IDX:
            col[LABEL_IDX["Profit/(loss) as per tally"]] = profit
        if "Difference" in LABEL_IDX:
            col[LABEL_IDX["Difference"]] = 0.0

        return col, profit, tot_alloc

    def _build_totals(agg: dict, pfx: str) -> list:
        """Build a totals column from an aggregation dict."""
        col: list = [0.0] * len(pl_rows)

        def _s(label: str, val: Any) -> None:
            if label in LABEL_IDX:
                col[LABEL_IDX[label]] = val

        sales  = agg["sales"]
        cogs   = agg["cogs"]
        dir_e  = agg["dir_exp"]
        gross  = agg["gross"]
        ind_inc = agg["ind_inc"]
        depr   = agg["depr"]
        tot_ind = agg["tot_ind"]
        profit  = agg["profit"]

        _s("Sales",   sales)
        _s("Less: COGS", cogs)
        if all_direct_ledgers:
            _s("3. Direct Expense", dir_e)
            for l in all_direct_ledgers:
                _s(l, agg["dir_breakdown"].get(l, 0.0))
        _s("Gross margin",  gross)
        _s("Gross margin %", _pct_str(gross, sales))
        _s("Gross margin % (previous month)", "0.00%")
        if all_income_ledgers:
            _s("Indirect Income", ind_inc)
            for l in all_income_ledgers:
                _s(l, agg["income_breakdown"].get(l, 0.0))
        else:
            _s("Indirect Income", ind_inc)
        _s("Depreciation & amortization", depr)
        _s("Net income",   ind_inc)
        _s("Net allocable income", 0.0)
        for l in all_indirect_ledgers:
            _s(l, agg["ind_breakdown"].get(l, 0.0))
        _s("Indirect costs", tot_ind)
        for cc in cost_centers:
            _s(cc, 0.0)
        _s("Total indirect costs", tot_ind)
        _s("Profit/ (loss) before tax", profit)
        _s("Net margin %", _pct_str(profit, sales))
        _s("Net margin % (previous month)", "0.00%")
        _s("Profit/(loss) as per tally", profit)
        _s("Difference", 0.0)
        return col

    # ------------------------------------------------------------------ #
    # 5.  Build pl_data dict: {col_name: [values]}                        #
    # ------------------------------------------------------------------ #
    pl_data: dict[str, list] = {"Particulars": pl_rows}

    def _zero_agg() -> dict:
        return {
            "sales": 0.0, "cogs": 0.0, "dir_exp": 0.0, "gross": 0.0,
            "ind_inc": 0.0, "depr": 0.0, "tot_ind": 0.0, "profit": 0.0,
            "ind_breakdown": {l: 0.0 for l in all_indirect_ledgers},
            "dir_breakdown": {l: 0.0 for l in all_direct_ledgers},
            "income_breakdown": {l: 0.0 for l in all_income_ledgers},
        }

    def _add_to_agg(agg: dict, nums: dict, profit: float) -> None:
        for k in ["sales", "cogs", "dir_exp", "gross", "ind_inc", "depr", "tot_ind"]:
            agg[k] += nums.get(k, 0.0)
        agg["profit"] += profit
        for l in all_indirect_ledgers:
            agg["ind_breakdown"][l] += nums["ind_breakdown"].get(l, 0.0)
        for l in all_direct_ledgers:
            agg["dir_breakdown"][l] += nums["dir_breakdown"].get(l, 0.0)
        for l in all_income_ledgers:
            agg["income_breakdown"][l] += nums["income_breakdown"].get(l, 0.0)

    for phase, pfx in [("CM", ""), ("YTD", "ytd_")]:
        agg_no_st  = _zero_agg()
        agg_inc_st = _zero_agg()

        for v in ordered_verts:
            vd = report_data["verticals"][v]
            col, nums = _build_col(vd, is_ytd=(phase == "YTD"))
            is_rev = vert_types.get(v) == "revenue"
            is_st  = vert_types.get(v) == "share_trading"

            col, profit, _ = _finalise_col(col, nums, v, is_rev, pfx)

            col_name = v if phase == "CM" else v + "_YTD"
            pl_data[col_name] = col

            if not is_st:
                _add_to_agg(agg_no_st, nums, profit)
            _add_to_agg(agg_inc_st, nums, profit)

        # --- Total columns ---
        tot_no_st_name  = "Total (without share trading)"  if phase == "CM" else "Total (without share trading)_YTD"
        tot_inc_st_name = "Total (including share trading)" if phase == "CM" else "Total (including share trading)_YTD"

        pl_data[tot_no_st_name]  = _build_totals(agg_no_st,  pfx)
        pl_data[tot_inc_st_name] = _build_totals(agg_inc_st, pfx)

        if phase == "CM":
            pl_data[" "] = [0.0] * len(pl_rows)   # spacer between CM and YTD

    pl_df = pd.DataFrame(pl_data)

    # ------------------------------------------------------------------ #
    # 6.  Drs_Crs data                                                    #
    # ------------------------------------------------------------------ #
    def _drcr_section(label: str, key: str) -> list[dict]:
        rows = [{"A": label, "B": "Opening", "C": "Debit", "D": "Credit", "E": "Closing",
                 "F": "", "G": "Opening", "H": "Debit", "I": "Credit", "J": "Closing"}]
        tots = {k: 0.0 for k in list("BCDE") + list("GHIJ")}
        for v in ordered_verts:
            d = report_data["verticals"].get(v, {}).get(key, {})
            if not any(d.values()):
                continue
            rows.append({
                "A": v, "B": d.get("opening", 0), "C": d.get("debit", 0),
                "D": d.get("credit", 0), "E": d.get("closing", 0), "F": "",
                "G": d.get("opening_ytd", 0), "H": d.get("debit_ytd", 0),
                "I": d.get("credit_ytd", 0), "J": d.get("closing_ytd", 0),
            })
            for k2, fld in zip("BCDE", ["opening", "debit", "credit", "closing"]):
                tots[k2] += d.get(fld, 0)
            for k2, fld in zip("GHIJ", ["opening_ytd", "debit_ytd", "credit_ytd", "closing_ytd"]):
                tots[k2] += d.get(fld, 0)
        rows.append({"A": "Grand Total", **{k: tots[k] for k in list("BCDE") + list("GHIJ")}, "F": ""})
        return rows

    drcr_rows = (
        _drcr_section("Sundry Debtor",  "debtors")
        + [{"A": "", "B": "", "C": "", "D": "", "E": "", "F": "", "G": "", "H": "", "I": "", "J": ""}]
        + _drcr_section("Sundry Creditor", "creditors")
    )
    drcr_df = pd.DataFrame(drcr_rows)

    # ------------------------------------------------------------------ #
    # 7.  Write to Excel                                                   #
    # ------------------------------------------------------------------ #
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        pl_df.to_excel(writer,   sheet_name="P&L",     index=False, startrow=6, header=False)
        drcr_df.to_excel(writer, sheet_name="Drs_Crs", index=False, startrow=4, header=False)

    wb = load_workbook(filepath)
    _format_pl(wb, pl_df, company_name, report_month, ytd_range,
               pl_rows, LABEL_IDX, all_direct_ledgers, all_income_ledgers,
               all_indirect_ledgers, cost_centers, ordered_verts, report_data)
    _format_drcr(wb, report_month, ytd_range, company_name)

    wb.save(filepath)
    return filename


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _format_pl(wb, pl_df, company_name, report_month, ytd_range,
               pl_rows, LABEL_IDX, all_direct_ledgers, all_income_ledgers,
               all_indirect_ledgers, cost_centers, ordered_verts, report_data):
    ws = wb["P&L"]

    # --- Meta rows ---
    ws["A1"].value = "M J P T & Co LLP";  ws["A1"].font = Font(bold=True, size=12)
    ws["A2"].value = company_name;         ws["A2"].font = Font(bold=True)
    ws["A4"].value = "P&L Analysis";      ws["A4"].font = Font(bold=True)
    ws["A5"].value = f"For the month of {report_month}"; ws["A5"].font = Font(bold=True)

    n_cols = len(pl_df.columns)
    col_names = list(pl_df.columns)

    # --- Compute CM / YTD boundary ---
    spacer_col = col_names.index(" ") + 1 if " " in col_names else None
    cm_total_col = None
    ytd_total_col = None
    for i, cn in enumerate(col_names):
        if cn == "Total (including share trading)":  cm_total_col  = i + 1
        if cn == "Total (including share trading)_YTD": ytd_total_col = i + 1

    if cm_total_col is None:
        for i, cn in enumerate(col_names):
            if cn == "Total (without share trading)": cm_total_col = i + 1
    if ytd_total_col is None:
        for i, cn in enumerate(col_names):
            if cn == "Total (without share trading)_YTD": ytd_total_col = i + 1

    # --- Month labels row 5 ---
    try:
        parts = report_month.split()
        short_mon = f"{parts[0]}'{parts[1][-2:]}"
    except Exception:
        short_mon = report_month

    if cm_total_col:
        c = ws.cell(row=5, column=cm_total_col, value=short_mon)
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    if spacer_col:
        ytd_start = spacer_col + 2
        c = ws.cell(row=5, column=ytd_start, value=ytd_range)
        c.font = Font(bold=True)
    if ytd_total_col:
        c = ws.cell(row=5, column=ytd_total_col, value=f"YTD'{report_month.split()[-1][-2:]}")
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")

    # --- Header row 6 ---
    ws.cell(row=6, column=1, value="Particulars").fill = _fill(DARK_BLUE)
    ws.cell(row=6, column=1).font = WHITE_FONT
    ws.cell(row=6, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for i, cn in enumerate(col_names):
        c_idx = i + 1
        display = cn.replace("_YTD", "").strip()
        cell = ws.cell(row=6, column=c_idx, value=display)
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c_idx)].width = 16
        if display == "":
            cell.fill = PatternFill(fill_type=None)
        elif "Total" in display or "Share Trading" in display:
            cell.fill = _fill(GREEN)
        else:
            cell.fill = _fill(DARK_BLUE)

    ws.column_dimensions["A"].width = 32

    # --- Data rows ---
    SECTION_HEADERS = {
        "Sales", "Less: COGS", "3. Direct Expense", "Gross margin",
        "Gross margin %", "Gross margin % (previous month)",
        "Indirect Income", "Depreciation & amortization",
        "Net income", "Net allocable income",
        "6. Indirect Expense", "Indirect costs",
        "Allocation of expenses:", "Total indirect costs",
        "Profit/ (loss) before tax", "Net margin %",
        "Net margin % (previous month)",
        "Profit/(loss) as per tally", "Difference",
    }
    BOLD_ROWS = {
        "Sales", "Less: COGS", "Gross margin", "Indirect Income",
        "Net income", "Indirect costs", "Total indirect costs",
        "Profit/ (loss) before tax", "Profit/(loss) as per tally",
    }
    BLUE_HEADER_ROWS = {
        "3. Direct Expense", "6. Indirect Expense", "Allocation of expenses:",
    }
    HIGHLIGHT_ROWS = {"Gross margin", "Net income", "Profit/ (loss) before tax"}
    PCT_ROWS = {
        "Gross margin %", "Gross margin % (previous month)",
        "Net margin %", "Net margin % (previous month)",
    }

    for row_offset, label in enumerate(pl_rows):
        excel_row = 7 + row_offset
        # Particulars cell
        a_cell = ws.cell(row=excel_row, column=1, value=label)
        if label in BOLD_ROWS:
            a_cell.font = Font(bold=True)
        if label in BLUE_HEADER_ROWS:
            a_cell.fill = _fill(LIGHT_BLUE)
            a_cell.font = Font(bold=True)
        if label in HIGHLIGHT_ROWS:
            a_cell.fill = _fill(YELLOW)
            a_cell.font = Font(bold=True)
        a_cell.alignment = Alignment(indent=0 if label in SECTION_HEADERS else 1)

        # Data cells
        for col_offset, cn in enumerate(col_names):
            c_idx  = col_offset + 1
            raw_val = pl_df.iloc[row_offset, col_offset]
            cell   = ws.cell(row=excel_row, column=c_idx)

            if label == "":
                cell.value = None
                continue

            if label in PCT_ROWS or isinstance(raw_val, str):
                # Keep percentage strings as-is
                cell.value = raw_val
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = float(raw_val) if raw_val not in (None, "") else 0.0
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")

            if label in BOLD_ROWS:
                cell.font = Font(bold=True)
            if label in HIGHLIGHT_ROWS:
                cell.fill = _fill(YELLOW)

    # --- Operating Expenses mini-table below main data ---
    last_row = ws.max_row
    summary_start = last_row + 2
    ws.cell(row=summary_start, column=1, value="Business Vertical").fill = _fill(DARK_BLUE)
    ws.cell(row=summary_start, column=1).font = WHITE_FONT
    ws.cell(row=summary_start, column=2, value=short_mon).fill = _fill(DARK_BLUE)
    ws.cell(row=summary_start, column=2).font = WHITE_FONT
    ws.cell(row=summary_start, column=3, value="YTD").fill = _fill(DARK_BLUE)
    ws.cell(row=summary_start, column=3).font = WHITE_FONT

    r = summary_start + 1
    for v in ordered_verts:
        cm_val  = sum(report_data["verticals"][v].get("indirect_breakdown", {}).values())
        ytd_val = sum(report_data["verticals"][v].get("ytd_indirect_breakdown", {}).values())
        if cm_val != 0 or ytd_val != 0:
            ws.cell(row=r, column=1, value=v)
            ws.cell(row=r, column=2, value=cm_val).number_format  = NUM_FMT
            ws.cell(row=r, column=3, value=ytd_val).number_format = NUM_FMT
            r += 1

    tot_cm  = sum(sum(report_data["verticals"][v].get("indirect_breakdown", {}).values())  for v in ordered_verts)
    tot_ytd = sum(sum(report_data["verticals"][v].get("ytd_indirect_breakdown", {}).values()) for v in ordered_verts)
    ws.cell(row=r, column=1, value="Total Operating Expenses").font = Font(bold=True)
    ws.cell(row=r, column=2, value=tot_cm).number_format  = NUM_FMT
    ws.cell(row=r, column=3, value=tot_ytd).number_format = NUM_FMT
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=3).font = Font(bold=True)


def _format_drcr(wb, report_month, ytd_range, company_name):
    ws = wb["Drs_Crs"]
    ws.cell(row=1, column=1, value="M J P T & Co LLP").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=company_name).font = Font(bold=True)
    ws.cell(row=4, column=1, value="Summary of Sundry Debtors & Sundry Creditors").font = Font(bold=True)

    try:
        parts = report_month.split()
        short_mon = f"{parts[0]}'{parts[1][-2:]}"
    except Exception:
        short_mon = report_month

    ws.cell(row=3, column=3, value=short_mon).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=6)
    ws.cell(row=3, column=8, value=ytd_range).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=11)

    ws.insert_cols(1)
    ws.column_dimensions["B"].width = 28
    for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[col_letter].width = 16

    # Apply number format to data cells
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=3, max_col=11):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FMT
